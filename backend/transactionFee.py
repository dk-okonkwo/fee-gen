from openpyxl import load_workbook
from datetime import datetime, timedelta

# 1. Unmerge Cells: prerequisite
# 2. Remove properties
# 3. Delete invoices outside date range
# 4. Delete "Payment Date", "Payment No.", "Unpaid", "Payee Type"
# 5. Delete bottom total row
# 6. Create "Trans Fee", "HST" and "Provinces" columns
# 7. Delete Wolverine transactions
# 8. Apply Schedule
# 9. Add HST


def get_month_range(group):
    '''
    Define the date range (from 11th of last month to 10th of this month)
    or 5th of last month to 4th of this month
    '''
    if group == "A":
        actual = 11
        today = datetime.today()
        start_date = today.replace(
            day=11) - timedelta(days=31)  # approx 11th
        diff = start_date.day - actual
        # correct to 11th of last month
        start = today.replace(day=11) - timedelta(days=(31+(diff)))
        end = today.replace(day=10)  # 10th of this month
        return start, end
    elif group == "B":
        actual = 5
        today = datetime.today()
        start_date = today.replace(
            day=5) - timedelta(days=31)  # approx 5th
        diff = start_date.day - actual
        # correct to 5th of last month
        start = today.replace(day=5) - timedelta(days=(31+(diff)))
        end = today.replace(day=4)  # 4th of this month

        return start, end


def get_transfee(group: str, trans_schedule: str, ledger: str):
    '''
    Calculates transaction fees
    '''
    wb = load_workbook(ledger)
    ws = wb.active

    prov = {
        'ON': 1.13,
        'NS': 1.15,
    }

    group_B = ['595 Silverbirch Road', '1100 Main Street West', '1117 Main Street West',
               '1098 Main Street West', '37 Highland Dr', '7 Foundry Street', '392 Albert St, Suite 304']

    start, end = get_month_range(group)

    sheet_data = 5
    wolverine = "Wolverine Property Management Limited"

    for row in reversed(range(sheet_data, ws.max_row + 1)):
        prop = ws.cell(row=row, column=6).value
        payee = ws.cell(row=row, column=2).value
        invoice_date = ws.cell(row=row, column=4).value

        # remove group B or A properties
        if isinstance(prop, str):
            if group == "A":
                if prop in group_B:
                    ws.delete_rows(row)
            elif group == "B":
                if prop not in group_B and row != 5:
                    ws.delete_rows(row)

        # remove wolverine transactions
        if isinstance(payee, str):
            if payee == wolverine:
                ws.delete_rows(row)

        # delete transactions outside date range
        if isinstance(invoice_date, str):
            try:
                # Convert string to datetime
                date_value = datetime.strptime(invoice_date, '%m/%d/%Y')
                # print(date_value)
                if start <= date_value <= end:
                    continue  # Date is within range, keep the row
                else:
                    # Date is not within range, delete the row
                    ws.delete_rows(row)
            except ValueError:
                print("not valid date")
                continue  # Skip if the date format is invalid
        else:
            print("not a string")

    for row in reversed(range(sheet_data, ws.max_row + 1)):
        ws.cell(row=row, column=1).value = None
        ws.cell(row=row, column=12).value = None

        # delete total row
        if ws.cell(row=row, column=7).value == "Grand Total":
            ws.delete_rows(row)

        # Add column headers
        if row == sheet_data:
            ws.cell(row=row, column=9).value = "Trans Fee"
            ws.cell(row=row, column=10).value = "+HST"
            ws.cell(row=row, column=11).value = "Province"
        else:
            ws.cell(row=row, column=9).value = None
            ws.cell(row=row, column=10).value = None
            ws.cell(row=row, column=11).value = None

    # calculate transaction Fee
    active_sch = load_workbook(trans_schedule)
    schedule = active_sch.active

    trans_fee = 9
    hst_col = 10
    location = 11

    for row in reversed(range(sheet_data + 1, ws.max_row + 1)):
        prop = ws.cell(row=row, column=6).value

        if prop:
            ledger_comma = ws.cell(row=row, column=8).value
            ledger_amount = ledger_comma
            if not isinstance(ledger_comma, float):
                ledger_str = ""
                for x in ledger_comma:
                    if x != ",":
                        ledger_str += x
                ledger_amount = float(ledger_str)
            for r in range(2, schedule.max_row + 1):
                if schedule.cell(row=r, column=1).value[:9] == prop[:9]:
                    loc = schedule.cell(row=r, column=7).value
                    ws.cell(row=row, column=11).value = loc
                    if ledger_amount < 11:
                        fee = schedule.cell(row=r, column=2).value
                        ws.cell(row=row, column=trans_fee).value = fee
                        ws.cell(row=row, column=hst_col).value = "{:.2f}".format(
                            int(fee) * prov[loc])
                        break
                    elif 11 <= ledger_amount < 19:
                        fee = schedule.cell(row=r, column=3).value
                        ws.cell(row=row, column=trans_fee).value = fee
                        ws.cell(row=row, column=hst_col).value = "{:.2f}".format(
                            int(fee) * prov[loc])
                        break
                    elif 19 <= ledger_amount < 1500:
                        mini = schedule.cell(row=r, column=3).value
                        per = float(schedule.cell(row=r, column=4).value)
                        per_ledger = ledger_amount * per
                        fee = max(per_ledger, mini)
                        ws.cell(row=row, column=trans_fee).value = fee
                        ws.cell(row=row, column=hst_col).value = "{:.2f}".format(
                            fee * prov[loc])
                        break
                    else:
                        maxi = schedule.cell(row=r, column=6).value
                        per = float(schedule.cell(row=r, column=5).value)
                        per_ledger = ledger_amount * per
                        fee = min(per_ledger, maxi)
                        ws.cell(row=row, column=trans_fee).value = fee
                        ws.cell(row=row, column=hst_col).value = "{:.2f}".format(
                            fee * prov[loc])
                if schedule.cell(row=r, column=1).value == "End":
                    break

    today = datetime.today()
    save_file_path = f'transactionFee{today.day}{today.minute}.xlsx'

    # Save the changes
    wb.save(f'sheets/{save_file_path}')

    return save_file_path
