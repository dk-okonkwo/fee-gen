from openpyxl import load_workbook
from datetime import datetime, timedelta

# 1. Find out month
# 2. Remove property not in group
# 3. Remove lease not in date range
# 4. If not "main tenant" remove row
# 5. Remove "phone" & "email" column
# 6. Create "Monthly Rent", "Leasing Fee", "+ HST", "Province"
# 7. Get tenant Rent
# 8. Leasing rate/amount
# 9. Get tax


def get_month_range(group):
    '''
    Define the date range (from 11th of last month to 10th of this month)
    or 5th of last month to 4th of this month
    '''
    if group == "A":
        actual = 11
        today = datetime.today()
        start_date = today.replace(
            day=11) - timedelta(days=31)  # approx 11th
        diff = start_date.day - actual
        # correct to 11th of last month
        start = today.replace(day=11) - timedelta(days=(31+(diff)))
        end = today.replace(day=10)  # 10th of this month
        return start, end

    elif group == "B":
        actual = 5
        today = datetime.today()
        start_date = today.replace(
            day=5) - timedelta(days=31)  # approx 5th
        diff = start_date.day - actual
        # correct to 5th of last month
        start = today.replace(day=5) - timedelta(days=(31+(diff)))
        end = today.replace(day=4)  # 4th of this month
        return start, end


def get_leasing_fee(group: str, tenants_rents_file: str, LARS_file: str, tenant_listing: str):
    # Load the Excel workbook
    wb = load_workbook(tenant_listing)

    # Select the worksheet
    ws = wb.active

    # Define the date range
    start, end = get_month_range(group)

    # Find the index of the date column
    lease_from = 7
    sheet_data = 7
    relationship = 5

    group_B = ['595 Silverbirch Road', '1100 Main Street West', '1117 Main Street West',
               '1098 Main Street West', '37 Highland Drive', '7 Foundry Street', '392 Albert Street, Suite 304']

    for row in reversed(range(sheet_data, ws.max_row + 1)):
        prop = ws.cell(row=row, column=1).value
        cell_value = ws.cell(row=row, column=lease_from).value
        tenant = ws.cell(row=row, column=relationship).value

        # remove group B or A properties
        if isinstance(prop, str):
            if group == "A":
                if prop in group_B:
                    ws.delete_rows(row)
            elif group == "B":
                if prop not in group_B:
                    ws.delete_rows(row)


        # remove non "main tenants"
        if tenant != "Main Tenant":
            ws.delete_rows(row)

        # remove properties not within date range
        if isinstance(cell_value, str):
            try:
                # Convert string to datetime
                date_value = datetime.strptime(cell_value, '%m/%d/%Y')
                # print(date_value)
                if start <= date_value <= end:
                    continue  # Date is within range, keep the row
                else:
                    # Date is not within range, delete the row
                    ws.delete_rows(row)
            except ValueError:
                print("not valid date")
                continue  # Skip if the date format is invalid
        else:
            print("not a string")

    # Remove "Phone" and "Email" columns
    phone_col = 10
    email_col = 11

    # Add column headers
    for row in reversed(range(6, ws.max_row + 1)):
        if row == 6:
            ws.cell(row=row, column=phone_col).value = "Monthly Rent"
            ws.cell(row=row, column=email_col).value = "Leasing Fee"
            ws.cell(row=row, column=email_col+1).value = "+HST"
            ws.cell(row=row, column=email_col+2).value = "Province"
        else:
            ws.cell(row=row, column=phone_col).value = None
            ws.cell(row=row, column=email_col).value = None

    prov = {
        'ON': 1.13,
        'NS': 1.15,
    }

    # Get Rent
    tenant_rents = load_workbook(tenants_rents_file)
    rents = tenant_rents.active

    for row in reversed(range(sheet_data, ws.max_row + 1)):
        tenant = ws.cell(row=row, column=3).value
        for r in reversed(range(5, rents.max_row + 1)):
            if rents.cell(row=r, column=2).value == tenant:
                ws.cell(row=row, column=phone_col).value = rents.cell(
                    row=r, column=3).value
                break

    ### GET LEASING RATE/AMOUNT AND TAX ###
    leasing_sch = load_workbook(LARS_file)
    lars = leasing_sch.active

    for row in range(sheet_data, ws.max_row + 1):
        prop = ws.cell(row=row, column=1).value
        rent_str = ws.cell(row=row, column=phone_col).value
        rent = int(rent_str) if rent_str else 0
        for r in range(2, lars.max_row + 1):
            if lars.cell(row=r, column=1).value == prop:
                percent = int(lars.cell(row=r, column=2).value)
                max_value = int(lars.cell(row=r, column=3).value)
                loc = lars.cell(row=r, column=6).value
                lease_amount = min((percent * rent), max_value)
                hst = "{:.2f}".format(lease_amount * prov[loc])
                ws.cell(row=row, column=email_col).value = lease_amount
                ws.cell(row=row, column=email_col+1).value = hst
                ws.cell(row=row, column=email_col+2).value = loc
            if lars.cell(row=r, column=1).value == None:
                break
        if ws.cell(row=row, column=1).value == "End":
            break

    today = datetime.today()
    save_file_path = f'leasingFee{today.day}{today.minute}.xlsx'

    # Save the changes
    wb.save(f'sheets/{save_file_path}')

    return save_file_path
