from openpyxl import load_workbook
from datetime import datetime


def get_tenants_rents(file: str):
    # file_date = datetime.now().strftime("%d-%m-%Y-%H%M%S")
    # new_file = 'tenantRents.xlsx'

    # Load the excel sheet
    wb = load_workbook(file)

    # select active sheet
    sheet = wb.active

    # iterate through rows in reversed order to avoid index isssues
    for row in reversed(range(5, sheet.max_row + 1)):
        # check that the second(tenant) column is empty
        prop = sheet.cell(row=row, column=2).value

        if not isinstance(prop, str):
            sheet.delete_rows(row)

    today = datetime.today()
    save_file_path = f'tenantRents{today.day}{today.minute}.xlsx'

    # Save the changes
    wb.save(f'sheets/{save_file_path}')

    return save_file_path