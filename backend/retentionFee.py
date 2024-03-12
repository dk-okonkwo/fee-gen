from openpyxl import load_workbook
from datetime import datetime, timedelta


# 1. Unmerge cells before uploading
# 2. Delete "Move In", "Move Out", "Notice Date", "Phone", "Alt Phone"
# 3. Find remove properties not in group
# 4. Remove properties that don't fall in the annivessary
# 5. Delete "Lease To" column
# 6. Remove all "Total"  and "Grand Total" row
# 7. Add "%", "Amount", "Fee to Charge(+ HST)" columns
# 8. If renewal percent in none then get the fixed amount
# 9. Get province and calculate + HST


def get_month_range(group):
    '''
    Define the date range (from 11th of last month to 10th of this month)
    or 5th of last month to 4th of this month
    '''
    if group == "A":
        actual = 11
        today = datetime.today()
        start_date = today.replace(
            day=11) - timedelta(days=31)  # approx 11th
        diff = start_date.day - actual
        # correct to 11th of last month
        start = today.replace(day=11) - timedelta(days=(31+(diff)))
        end = today.replace(day=10)  # 10th of this month
        return start, end

    elif group == "B":
        actual = 5
        today = datetime.today()
        start_date = today.replace(
            day=5) - timedelta(days=31)  # approx 5th
        diff = start_date.day - actual
        # correct to 5th of last month
        start = today.replace(day=5) - timedelta(days=(31+(diff)))
        end = today.replace(day=4)  # 4th of this month
        return start, end


def get_renewal_fee(group: str, LARS_file: str, tenant_activity: str):
    '''
    Runs through the 9 steps above to generate retention fee excel sheet
    '''

    group_B = ['595 Silverbirch Road', '1100 Main Street West', '1117 Main Street West',
               '1098 Main Street West', '37 Highland Dr', '7 Foundry Street', '392 Albert St, Suite 304']

    wb = load_workbook(tenant_activity)
    ws = wb.active

    lease_to_index = 7
    email_index = 13
    sheet_data = 4
    last_total = 0
    lease_from_index = 6

    prov = {
        'ON': 1.13,
        'NS': 1.15,
    }

    start, end = get_month_range(group)

    for row in reversed(range(sheet_data, ws.max_row + 1)):
        # remove unnecessary columns
        for i in range(lease_to_index, email_index):
            ws.cell(row=row, column=i).value = None

        # Remove group B or A properties
        prop = ws.cell(row=row, column=1).value
        col_2 = ws.cell(row=row, column=2).value
        if prop == "Total ":
            last_total = row
        if group == "A":
            if prop in group_B:
                # print(last_total)
                while last_total >= row:
                    ws.delete_rows(last_total)
                    last_total -= 1
        elif group == "B":
            if prop not in group_B and col_2 is None and prop != "Total ":
                # print(last_total)
                while last_total >= row:
                    ws.delete_rows(last_total)
                    last_total -= 1

    # remove the remaining total rows
    for row in reversed(range(sheet_data, ws.max_row + 1)):
        lease = ws.cell(row=row, column=lease_from_index).value
        prop = ws.cell(row=row, column=1).value
        if prop == "Total " or prop == "Grand Total ":
            ws.delete_rows(row)

        # Add column headers
        if row == 4:
            ws.cell(row=row, column=lease_to_index).value = "Amount"
            ws.cell(row=row, column=lease_to_index +
                    1).value = "Fee to charge (+HST)"
            ws.cell(row=row, column=lease_to_index+2).value = "Province"

        if isinstance(lease, datetime):
            if lease.year == end.year:
                ws.delete_rows(row)
            elif (lease.month == start.month) and (lease.year == start.year):
                ws.delete_rows(row)
            elif (start.month == lease.month) and (start.day <= lease.day):
                continue
            elif (end.month == lease.month) and (end.day >= lease.day):
                continue
            else:
                ws.delete_rows(row)

    count = 0
    # remove excees properties
    for row in reversed(range(5, ws.max_row + 1)):
        if ws.cell(row=row, column=6).value is None:
            count += 1
        if ws.cell(row=row, column=6).value is not None:
            count = 0
        if count > 1:
            ws.delete_rows(row)
            count -= 1

    ### GET RETENTION RATE/AMOUNT AND TAX ###
    leasing_sch = load_workbook(LARS_file)
    lars = leasing_sch.active
    fee = 7
    hst = 8
    filled = 0
    amount = 0
    percent = None
    loc = None
    for row in reversed(range(5, ws.max_row + 1)):
        if ws.cell(row=row, column=6).value is not None:
            filled += 1
        if filled > 0 and ws.cell(row=row, column=6).value is None:
            prop = ws.cell(row=row, column=1).value
            for r in range(2, lars.max_row + 1):
                if lars.cell(row=r, column=1).value[:9] == prop[:9]:
                    loc = lars.cell(row=r, column=6).value
                    ws.cell(row=row, column=9).value = loc
                    percent = lars.cell(row=r, column=4).value
                    fixed = lars.cell(row=r, column=5).value
                    amount = float(percent) if percent else int(fixed)
                    for unit in range(1, filled+1):
                        ws.cell(row=row+unit, column=9).value = loc
                        if percent:
                            rent = ws.cell(row=row+unit, column=3).value
                            ws.cell(row=row+unit,
                                    column=fee).value = float(rent) * amount
                            ws.cell(
                                row=row+unit, column=hst).value = "{:.2f}".format(float(rent) * amount * prov[loc])
                        else:
                            ws.cell(row=row+unit, column=fee).value = amount
                            ws.cell(
                                row=row+unit, column=hst).value = "{:.2f}".format(amount * prov[loc])
                    break
            filled = 0

    today = datetime.today()
    save_file_path = f'retentionFee{today.day}{today.minute}.xlsx'

    # Save the changes
    wb.save(f'sheets/{save_file_path}')

    return save_file_path